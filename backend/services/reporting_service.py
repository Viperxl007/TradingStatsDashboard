"""
Data Export and Reporting Service for CL Position Tracking

This module provides comprehensive reporting capabilities including:
- PDF report generation
- Excel export with advanced formatting
- Tax reporting and P&L statements
- Performance benchmarking reports
- Custom report templates and scheduling
- API for third-party integrations
"""

import logging
import sqlite3
from datetime import datetime, timedelta
from typing import Dict, Any, Optional, List, Union
from dataclasses import dataclass
from enum import Enum
import json
import uuid
from threading import Lock
import io
import base64

# PDF generation
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.linecharts import HorizontalLineChart
from reportlab.graphics.charts.piecharts import Pie

# Excel generation
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class ReportType(Enum):
    """Report type enumeration."""
    PORTFOLIO_SUMMARY = "portfolio_summary"
    POSITION_DETAIL = "position_detail"
    PERFORMANCE_ANALYSIS = "performance_analysis"
    TAX_REPORT = "tax_report"
    RISK_ANALYSIS = "risk_analysis"
    CUSTOM = "custom"


class ExportFormat(Enum):
    """Export format enumeration."""
    PDF = "pdf"
    EXCEL = "excel"
    CSV = "csv"
    JSON = "json"


@dataclass
class ReportConfig:
    """Report configuration data structure."""
    id: str
    name: str
    report_type: ReportType
    export_format: ExportFormat
    parameters: Dict[str, Any]
    template_id: Optional[str] = None
    schedule: Optional[str] = None
    recipients: List[str] = None
    created_at: datetime = None
    updated_at: datetime = None


@dataclass
class ReportTemplate:
    """Report template data structure."""
    id: str
    name: str
    report_type: ReportType
    template_data: Dict[str, Any]
    is_default: bool = False
    created_at: datetime = None


class ReportingService:
    """
    Comprehensive reporting service for CL position tracking.
    
    Provides advanced report generation, export capabilities,
    and scheduled reporting functionality.
    """
    
    def __init__(self, db_path: Optional[str] = None):
        """
        Initialize the reporting service.
        
        Args:
            db_path (Optional[str]): Path to SQLite database file
        """
        import os
        self.db_path = db_path or os.path.join(
            os.path.dirname(__file__), '..', 'instance', 'reports.db'
        )
        self.db_lock = Lock()
        
        # Initialize database
        self._ensure_database()
        
        # Initialize default templates
        self._create_default_templates()
        
        logger.info("Reporting Service initialized")
    
    def _ensure_database(self):
        """Ensure the database and tables exist."""
        try:
            import os
            os.makedirs(os.path.dirname(self.db_path), exist_ok=True)
            
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                
                # Create report_configs table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS report_configs (
                        id TEXT PRIMARY KEY,
                        name TEXT NOT NULL,
                        report_type TEXT NOT NULL,
                        export_format TEXT NOT NULL,
                        parameters TEXT NOT NULL,
                        template_id TEXT,
                        schedule TEXT,
                        recipients TEXT,
                        created_at INTEGER NOT NULL,
                        updated_at INTEGER NOT NULL
                    )
                ''')
                
                # Create report_templates table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS report_templates (
                        id TEXT PRIMARY KEY,
                        name TEXT NOT NULL,
                        report_type TEXT NOT NULL,
                        template_data TEXT NOT NULL,
                        is_default INTEGER DEFAULT 0,
                        created_at INTEGER NOT NULL
                    )
                ''')
                
                # Create report_history table
                cursor.execute('''
                    CREATE TABLE IF NOT EXISTS report_history (
                        id TEXT PRIMARY KEY,
                        config_id TEXT,
                        report_type TEXT NOT NULL,
                        export_format TEXT NOT NULL,
                        file_path TEXT,
                        file_size INTEGER,
                        generation_time_ms INTEGER,
                        status TEXT DEFAULT 'completed',
                        error_message TEXT,
                        created_at INTEGER NOT NULL,
                        FOREIGN KEY (config_id) REFERENCES report_configs (id)
                    )
                ''')
                
                # Create indexes
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_configs_type ON report_configs(report_type)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_templates_type ON report_templates(report_type)')
                cursor.execute('CREATE INDEX IF NOT EXISTS idx_history_created ON report_history(created_at)')
                
                conn.commit()
                logger.info(f"Reporting database initialized at {self.db_path}")
                
        except Exception as e:
            logger.error(f"Error initializing reporting database: {str(e)}")
            raise
    
    def generate_portfolio_summary_pdf(self, positions: List[Dict[str, Any]], 
                                     portfolio_metrics: Dict[str, Any]) -> bytes:
        """
        Generate a comprehensive portfolio summary PDF report.
        
        Args:
            positions (List[Dict[str, Any]]): List of positions
            portfolio_metrics (Dict[str, Any]): Portfolio metrics
            
        Returns:
            bytes: PDF file content
        """
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph("Concentrated Liquidity Portfolio Report", title_style))
            story.append(Spacer(1, 20))
            
            # Report metadata
            report_date = datetime.now().strftime("%B %d, %Y")
            story.append(Paragraph(f"<b>Report Date:</b> {report_date}", styles['Normal']))
            story.append(Paragraph(f"<b>Total Positions:</b> {len(positions)}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Executive Summary
            story.append(Paragraph("Executive Summary", styles['Heading2']))
            
            summary_data = [
                ['Metric', 'Value'],
                ['Total Portfolio Value', f"${portfolio_metrics.get('current_value', 0):,.2f}"],
                ['Total Investment', f"${portfolio_metrics.get('total_investment', 0):,.2f}"],
                ['Total Fees Collected', f"${portfolio_metrics.get('total_fees_collected', 0):,.2f}"],
                ['Total P&L', f"${portfolio_metrics.get('total_pnl', 0):,.2f}"],
                ['Total Return %', f"{portfolio_metrics.get('total_return_pct', 0):.2f}%"],
                ['Active Positions', str(portfolio_metrics.get('active_positions', 0))],
                ['Closed Positions', str(portfolio_metrics.get('closed_positions', 0))]
            ]
            
            summary_table = Table(summary_data, colWidths=[2.5*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(summary_table)
            story.append(Spacer(1, 30))
            
            # Position Details
            story.append(Paragraph("Position Details", styles['Heading2']))
            
            if positions:
                position_data = [['Position', 'Pair', 'Status', 'Investment', 'Current Value', 'Fees', 'P&L', 'Return %']]
                
                for pos in positions:
                    current_value = pos.get('current_value', pos['initial_investment'])
                    fees = pos.get('fees_collected', 0)
                    pnl = current_value + fees - pos['initial_investment']
                    return_pct = (pnl / pos['initial_investment'] * 100) if pos['initial_investment'] > 0 else 0
                    
                    position_data.append([
                        pos['trade_name'][:20],  # Truncate long names
                        pos['pair_symbol'],
                        pos['status'].title(),
                        f"${pos['initial_investment']:,.0f}",
                        f"${current_value:,.0f}",
                        f"${fees:,.0f}",
                        f"${pnl:,.0f}",
                        f"{return_pct:.1f}%"
                    ])
                
                position_table = Table(position_data, colWidths=[1.2*inch, 0.8*inch, 0.7*inch, 0.8*inch, 0.8*inch, 0.6*inch, 0.7*inch, 0.6*inch])
                position_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(position_table)
            else:
                story.append(Paragraph("No positions found.", styles['Normal']))
            
            story.append(Spacer(1, 30))
            
            # Performance Analysis
            story.append(Paragraph("Performance Analysis", styles['Heading2']))
            
            best_position = portfolio_metrics.get('best_performing_position', 'N/A')
            worst_position = portfolio_metrics.get('worst_performing_position', 'N/A')
            
            story.append(Paragraph(f"<b>Best Performing Position:</b> {best_position}", styles['Normal']))
            story.append(Paragraph(f"<b>Worst Performing Position:</b> {worst_position}", styles['Normal']))
            story.append(Spacer(1, 10))
            
            # Risk Metrics (if available)
            if 'risk_metrics' in portfolio_metrics:
                risk_metrics = portfolio_metrics['risk_metrics']
                story.append(Paragraph("Risk Metrics", styles['Heading3']))
                story.append(Paragraph(f"<b>Maximum Drawdown:</b> {risk_metrics.get('max_drawdown', 0):.2%}", styles['Normal']))
                story.append(Paragraph(f"<b>Sharpe Ratio:</b> {risk_metrics.get('sharpe_ratio', 0):.2f}", styles['Normal']))
                story.append(Paragraph(f"<b>Volatility:</b> {risk_metrics.get('volatility', 0):.2%}", styles['Normal']))
            
            # Footer
            story.append(Spacer(1, 50))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                alignment=1
            )
            story.append(Paragraph("Generated by CL Position Tracking System", footer_style))
            story.append(Paragraph(f"Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))
            
            # Build PDF
            doc.build(story)
            buffer.seek(0)
            
            logger.info("Generated portfolio summary PDF report")
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating portfolio summary PDF: {str(e)}")
            raise
    
    def generate_portfolio_summary_excel(self, positions: List[Dict[str, Any]], 
                                       portfolio_metrics: Dict[str, Any]) -> bytes:
        """
        Generate a comprehensive portfolio summary Excel report.
        
        Args:
            positions (List[Dict[str, Any]]): List of positions
            portfolio_metrics (Dict[str, Any]): Portfolio metrics
            
        Returns:
            bytes: Excel file content
        """
        try:
            buffer = io.BytesIO()
            
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                # Portfolio Summary Sheet
                summary_data = {
                    'Metric': [
                        'Total Portfolio Value',
                        'Total Investment',
                        'Total Fees Collected',
                        'Total P&L',
                        'Total Return %',
                        'Active Positions',
                        'Closed Positions'
                    ],
                    'Value': [
                        portfolio_metrics.get('current_value', 0),
                        portfolio_metrics.get('total_investment', 0),
                        portfolio_metrics.get('total_fees_collected', 0),
                        portfolio_metrics.get('total_pnl', 0),
                        portfolio_metrics.get('total_return_pct', 0),
                        portfolio_metrics.get('active_positions', 0),
                        portfolio_metrics.get('closed_positions', 0)
                    ]
                }
                
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Portfolio Summary', index=False)
                
                # Position Details Sheet
                if positions:
                    position_data = []
                    for pos in positions:
                        current_value = pos.get('current_value', pos['initial_investment'])
                        fees = pos.get('fees_collected', 0)
                        pnl = current_value + fees - pos['initial_investment']
                        return_pct = (pnl / pos['initial_investment'] * 100) if pos['initial_investment'] > 0 else 0
                        
                        position_data.append({
                            'Position Name': pos['trade_name'],
                            'Pair Symbol': pos['pair_symbol'],
                            'Status': pos['status'].title(),
                            'Protocol': pos.get('protocol', 'Unknown'),
                            'Entry Date': pos['entry_date'],
                            'Exit Date': pos.get('exit_date', ''),
                            'Price Range Min': pos['price_range_min'],
                            'Price Range Max': pos['price_range_max'],
                            'Initial Investment': pos['initial_investment'],
                            'Current Value': current_value,
                            'Fees Collected': fees,
                            'P&L': pnl,
                            'Return %': return_pct,
                            'APR %': pos.get('apr', 0),
                            'Current Price': pos.get('current_price', ''),
                            'In Range': pos.get('is_in_range', ''),
                            'Notes': pos.get('notes', '')
                        })
                    
                    positions_df = pd.DataFrame(position_data)
                    positions_df.to_excel(writer, sheet_name='Position Details', index=False)
                
                # Performance Analysis Sheet
                perf_data = {
                    'Analysis': [
                        'Best Performing Position',
                        'Worst Performing Position',
                        'Average Position Return',
                        'Win Rate',
                        'Total Trades'
                    ],
                    'Value': [
                        portfolio_metrics.get('best_performing_position', 'N/A'),
                        portfolio_metrics.get('worst_performing_position', 'N/A'),
                        f"{portfolio_metrics.get('avg_return_pct', 0):.2f}%",
                        f"{portfolio_metrics.get('win_rate', 0):.1f}%",
                        len(positions)
                    ]
                }
                
                perf_df = pd.DataFrame(perf_data)
                perf_df.to_excel(writer, sheet_name='Performance Analysis', index=False)
                
                # Format the Excel file
                workbook = writer.book
                
                # Format Portfolio Summary sheet
                summary_sheet = workbook['Portfolio Summary']
                self._format_excel_sheet(summary_sheet, 'Portfolio Summary')
                
                # Format Position Details sheet
                if 'Position Details' in workbook.sheetnames:
                    details_sheet = workbook['Position Details']
                    self._format_excel_sheet(details_sheet, 'Position Details')
                
                # Format Performance Analysis sheet
                perf_sheet = workbook['Performance Analysis']
                self._format_excel_sheet(perf_sheet, 'Performance Analysis')
            
            buffer.seek(0)
            
            logger.info("Generated portfolio summary Excel report")
            return buffer.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating portfolio summary Excel: {str(e)}")
            raise
    
    def generate_tax_report(self, positions: List[Dict[str, Any]], 
                          tax_year: int) -> Dict[str, Any]:
        """
        Generate tax report for a specific year.
        
        Args:
            positions (List[Dict[str, Any]]): List of positions
            tax_year (int): Tax year
            
        Returns:
            Dict[str, Any]: Tax report data
        """
        try:
            # Filter positions for the tax year
            year_start = datetime(tax_year, 1, 1)
            year_end = datetime(tax_year, 12, 31)
            
            relevant_positions = []
            for pos in positions:
                entry_date = datetime.fromisoformat(pos['entry_date'].replace('Z', '+00:00'))
                exit_date = None
                
                if pos.get('exit_date'):
                    exit_date = datetime.fromisoformat(pos['exit_date'].replace('Z', '+00:00'))
                
                # Include if opened or closed in tax year
                if (year_start <= entry_date <= year_end) or (exit_date and year_start <= exit_date <= year_end):
                    relevant_positions.append(pos)
            
            # Calculate tax implications
            total_fees_income = 0
            realized_gains = 0
            realized_losses = 0
            unrealized_gains = 0
            
            taxable_events = []
            
            for pos in relevant_positions:
                fees = pos.get('fees_collected', 0)
                total_fees_income += fees
                
                # Add fee income as taxable event
                if fees > 0:
                    taxable_events.append({
                        'date': pos.get('exit_date', pos['entry_date']),
                        'type': 'Fee Income',
                        'position': pos['trade_name'],
                        'amount': fees,
                        'description': f"Liquidity provision fees for {pos['pair_symbol']}"
                    })
                
                # Calculate gains/losses for closed positions
                if pos['status'] == 'closed' and pos.get('exit_date'):
                    exit_date = datetime.fromisoformat(pos['exit_date'].replace('Z', '+00:00'))
                    
                    if year_start <= exit_date <= year_end:
                        current_value = pos.get('current_value', pos['initial_investment'])
                        capital_gain = current_value - pos['initial_investment']
                        
                        if capital_gain > 0:
                            realized_gains += capital_gain
                        else:
                            realized_losses += abs(capital_gain)
                        
                        taxable_events.append({
                            'date': pos['exit_date'],
                            'type': 'Capital Gain/Loss',
                            'position': pos['trade_name'],
                            'amount': capital_gain,
                            'description': f"Position closure for {pos['pair_symbol']}"
                        })
                
                # Calculate unrealized gains for open positions
                elif pos['status'] == 'active':
                    current_value = pos.get('current_value', pos['initial_investment'])
                    unrealized_gain = current_value - pos['initial_investment']
                    unrealized_gains += unrealized_gain
            
            tax_report = {
                'tax_year': tax_year,
                'summary': {
                    'total_fee_income': total_fees_income,
                    'realized_capital_gains': realized_gains,
                    'realized_capital_losses': realized_losses,
                    'net_realized_gains': realized_gains - realized_losses,
                    'unrealized_gains': unrealized_gains,
                    'total_taxable_income': total_fees_income + realized_gains - realized_losses
                },
                'taxable_events': sorted(taxable_events, key=lambda x: x['date']),
                'positions_analyzed': len(relevant_positions),
                'generated_at': datetime.now().isoformat(),
                'disclaimer': "This report is for informational purposes only. Consult a tax professional for official tax advice."
            }
            
            logger.info(f"Generated tax report for year {tax_year}")
            return tax_report
            
        except Exception as e:
            logger.error(f"Error generating tax report: {str(e)}")
            raise
    
    def export_positions_csv(self, positions: List[Dict[str, Any]]) -> str:
        """
        Export positions to CSV format.
        
        Args:
            positions (List[Dict[str, Any]]): List of positions
            
        Returns:
            str: CSV content
        """
        try:
            if not positions:
                return "No positions to export"
            
            # Prepare data for CSV
            csv_data = []
            headers = [
                'Position Name', 'Pair Symbol', 'Status', 'Protocol', 'Chain',
                'Entry Date', 'Exit Date', 'Price Range Min', 'Price Range Max',
                'Initial Investment', 'Current Value', 'Fees Collected', 'P&L',
                'Return %', 'APR %', 'Current Price', 'In Range', 'Notes'
            ]
            
            csv_data.append(','.join(headers))
            
            for pos in positions:
                current_value = pos.get('current_value', pos['initial_investment'])
                fees = pos.get('fees_collected', 0)
                pnl = current_value + fees - pos['initial_investment']
                return_pct = (pnl / pos['initial_investment'] * 100) if pos['initial_investment'] > 0 else 0
                
                row = [
                    f'"{pos["trade_name"]}"',
                    pos['pair_symbol'],
                    pos['status'],
                    pos.get('protocol', ''),
                    pos.get('chain', ''),
                    pos['entry_date'],
                    pos.get('exit_date', ''),
                    str(pos['price_range_min']),
                    str(pos['price_range_max']),
                    str(pos['initial_investment']),
                    str(current_value),
                    str(fees),
                    str(pnl),
                    f"{return_pct:.2f}",
                    str(pos.get('apr', 0)),
                    str(pos.get('current_price', '')),
                    str(pos.get('is_in_range', '')),
                    f'"{pos.get("notes", "")}"'
                ]
                
                csv_data.append(','.join(row))
            
            csv_content = '\n'.join(csv_data)
            
            logger.info(f"Exported {len(positions)} positions to CSV")
            return csv_content
            
        except Exception as e:
            logger.error(f"Error exporting positions to CSV: {str(e)}")
            raise
    
    def _format_excel_sheet(self, sheet, sheet_name: str):
        """Format Excel sheet with styling."""
        try:
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Apply header formatting to first row
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in sheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
            
        except Exception as e:
            logger.error(f"Error formatting Excel sheet {sheet_name}: {str(e)}")
    
    def _create_default_templates(self):
        """Create default report templates."""
        try:
            templates = [
                {
                    'name': 'Standard Portfolio Summary',
                    'report_type': ReportType.PORTFOLIO_SUMMARY,
                    'template_data': {
                        'sections': ['executive_summary', 'position_details', 'performance_analysis'],
                        'charts': ['portfolio_allocation', 'performance_trend'],
                        'styling': 'professional'
                    },
                    'is_default': True
                },
                {
                    'name': 'Detailed Position Report',
                    'report_type': ReportType.POSITION_DETAIL,
                    'template_data': {
                        'sections': ['position_overview', 'price_history', 'fee_analysis', 'risk_metrics'],
                        'charts': ['price_chart', 'fee_collection_trend'],
                        'styling': 'detailed'
                    },
                    'is_default': True
                },
                {
                    'name': 'Tax Report Template',
                    'report_type': ReportType.TAX_REPORT,
                    'template_data': {
                        'sections': ['tax_summary', 'taxable_events', 'position_breakdown'],
                        'include_disclaimers': True,
                        'styling': 'formal'
                    },
                    'is_default': True
                }
            ]
            
            for template_data in templates:
                # Check if template already exists
                with sqlite3.connect(self.db_path) as conn:
                    cursor = conn.cursor()
                    cursor.execute(
                        'SELECT id FROM report_templates WHERE name = ? AND report_type = ?',
                        (template_data['name'], template_data['report_type'].value)
                    )
                    
                    if not cursor.fetchone():
                        # Create new template
                        template_id = str(uuid.uuid4())
                        current_timestamp = int(datetime.now().timestamp())
                        
                        cursor.execute('''
                            INSERT INTO report_templates (
                                id, name, report_type, template_data, is_default, created_at
                            ) VALUES (?, ?, ?, ?, ?, ?)
                        ''', (
                            template_id, template_data['name'], template_data['report_type'].value,
                            json.dumps(template_data['template_data']), int(template_data['is_default']),
                            current_timestamp
                        ))
                        
                        conn.commit()
                        logger.info(f"Created default template: {template_data['name']}")
            
        except Exception as e:
            logger.error(f"Error creating default templates: {str(e)}")
    
    def save_report_config(self, config: ReportConfig) -> str:
        """
        Save a report configuration.
        
        Args:
            config (ReportConfig): Report configuration
            
        Returns:
            str: Configuration ID
        """
        try:
            config_id = config.id or str(uuid.uuid4())
            current_timestamp = int(datetime.now().timestamp())
            
            with self.db_lock:
                with sqlite3.connect(self.db_path) as conn:
                    cursor = conn.cursor()
                    
                    cursor.execute('''
                        INSERT OR REPLACE INTO report_configs (
                            id, name, report_type, export_format, parameters,
                            template_id, schedule, recipients, created_at, updated_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        config_id, config.name, config.report_type.value, config.export_format.value,
                        json.dumps(config.parameters), config.template_id, config.schedule,
                        json.dumps(config.recipients) if config.recipients else None,
                        current_timestamp, current_timestamp
                    ))
                    
                    conn.commit()
                    
            logger.info(f"Saved report configuration: {config_id}")
            return config_id
            
        except Exception as e:
            logger.error(f"Error saving report configuration: {str(e)}")
            raise
    
    def get_report_configs(self) -> List[Dict[str, Any]]:
        """
        Get all report configurations.
        
        Returns:
            List[Dict[str, Any]]: List of report configurations
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                cursor.execute('SELECT * FROM report_configs ORDER BY created_at DESC')
                rows = cursor.fetchall()
                
                configs = []
                for row in rows:
                    config_dict = dict(row)
                    # Parse JSON fields
                    config_dict['parameters'] = json.loads(config_dict['parameters'])
                    if config_dict['recipients']:
                        config_dict['recipients'] = json.loads(config_dict['recipients'])
                    configs.append(config_dict)
                
                return configs
                
        except Exception as e:
            logger.error(f"Error retrieving report configurations: {str(e)}")
            raise
    
    def get_report_history(self, limit: Optional[int] = None) -> List[Dict[str, Any]]:
        """
        Get report generation history.
        
        Args:
            limit (Optional[int]): Limit number of results
            
        Returns:
            List[Dict[str, Any]]: List of report history records
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                conn.row_factory = sqlite3.Row
                cursor = conn.cursor()
                
                query = 'SELECT * FROM report_history ORDER BY created_at DESC'
                params = []
                
                if limit:
                    query += ' LIMIT ?'
                    params.append(limit)
                
                cursor.execute(query, params)
                rows = cursor.fetchall()
                
                return [dict(row) for row in rows]
                
        except Exception as e:
            logger.error(f"Error retrieving report history: {str(e)}")
            raise
    
    def log_report_generation(self, report_type: str, export_format: str,
                            file_path: Optional[str] = None, file_size: Optional[int] = None,
                            generation_time_ms: Optional[int] = None,
                            status: str = 'completed', error_message: Optional[str] = None) -> str:
        """
        Log report generation activity.
        
        Args:
            report_type (str): Type of report generated
            export_format (str): Export format used
            file_path (Optional[str]): Path to generated file
            file_size (Optional[int]): Size of generated file in bytes
            generation_time_ms (Optional[int]): Generation time in milliseconds
            status (str): Generation status
            error_message (Optional[str]): Error message if failed
            
        Returns:
            str: History record ID
        """
        try:
            history_id = str(uuid.uuid4())
            current_timestamp = int(datetime.now().timestamp())
            
            with self.db_lock:
                with sqlite3.connect(self.db_path) as conn:
                    cursor = conn.cursor()
                    
                    cursor.execute('''
                        INSERT INTO report_history (
                            id, report_type, export_format, file_path, file_size,
                            generation_time_ms, status, error_message, created_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (
                        history_id, report_type, export_format, file_path, file_size,
                        generation_time_ms, status, error_message, current_timestamp
                    ))
                    
                    conn.commit()
                    
            return history_id
            
        except Exception as e:
            logger.error(f"Error logging report generation: {str(e)}")
            raise